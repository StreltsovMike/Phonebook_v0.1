
import openpyxl as op

def read_xlsx():   # Выдает список с значениями каждой ячейки из phoonebook.xlsx
    data = op.open('phonebook.xlsx', read_only=True)
    sheet = data.active
    data_list = []
    for row in range(1, (sheet.max_row + 1)):
        for col in range(5):
            data_list.append(sheet[row][col].value)
    return data_list

def updated_txt_v1(my_list):   # Создает phonebook_v1.txt в котором каждая ячейка расположена через пустую строку построчно
    data = open('phonebook_v1.txt', 'w')
    columns = ['id: ', 'Name: ', 'Surname: ', 'Phone Number: ', 'Description: ']
    for i in range(5, len(my_list)):
        if isinstance(my_list[i], float):
            my_list[i] = int(my_list[i])
        data.write(f'\n{my_list[i]}\n')

def updated_txt_v2(my_list):   # Создает phonebook_v2.txt в котором каждый номер и его атрибуты расположены построчно
    data = open('phonebook_v2.txt', 'w')
    for i in range(len(my_list)):
        if isinstance(my_list[i], float):
            my_list[i] = int(my_list[i])
        if i % 5 == 0 and i != 0:
            data.write('\n')
        data.write(f'{my_list[i]},   ')

def add_contact_to_list(new_person):  # Добавляет в список контактов новый контак, чтобы дальше перезаписать файл
    data_list  = read_xlsx()
    last_id = int(data_list[len(data_list)-5])
    new_person.insert(0, last_id + 1 )
    for i in range(len(new_person)):
        data_list.append(new_person[i])
    print(data_list[0], data_list[1])
    return data_list


def add_contact_to_exel(my_list):  # Перезаписывает файл .xlsx со всеми изменениями
    data = op.Workbook()
    sheet = data.active
    row = 1
    count = 1 
    column = 1
    for i in range(1, len(my_list)+1):
        if count == 6:
            row +=1
            column = 1
            count = 1
        sheet.cell(row = row, column = column).value = my_list[i-1]
        count += 1
        column +=1
    data.save('phonebook.xlsx')
    data.close

def find_name(name):    # Находит контакты по указанному имени и выводит в терминал
    data = op.open('phonebook.xlsx', read_only=True)
    sheet = data.active
    contacts_list = []
    for i in range(1, sheet.max_row):
        if str(sheet.cell(row = i, column = 2).value) == name:
            for j in range(5):
                contacts_list.append(sheet[i][j].value)
    print(contacts_list)
    return contacts_list

# name = 'Mary'
# find_name(name)